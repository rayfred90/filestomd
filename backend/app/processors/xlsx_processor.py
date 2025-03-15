from typing import Dict, Any, BinaryIO, List
import openpyxl
from openpyxl.utils import get_column_letter
from .base_processor import BaseProcessor
from ..utils.chunker import DocumentChunker

class XlsxProcessor(BaseProcessor):
    def __init__(self):
        super().__init__()
        self.chunker = DocumentChunker(max_chunk_size=1000)  # Smaller chunks for table data

    def process(self, file: BinaryIO) -> Dict[str, Any]:
        """Process an XLSX file and extract its content with metadata."""
        # Load workbook
        workbook = openpyxl.load_workbook(file, data_only=True)  # data_only=True to get values instead of formulas
        
        # Extract content and metadata for each sheet
        sheets_data = []
        all_content = []
        
        for sheet in workbook.worksheets:
            sheet_data = self._process_sheet(sheet)
            sheets_data.append(sheet_data)
            all_content.append(f"# {sheet.title}\n\n{sheet_data['content']}")
        
        # Combine all content
        full_content = "\n\n".join(all_content)
        
        # Extract overall metadata
        metadata = self._extract_metadata(workbook, sheets_data)
        
        # Generate chunks (by sheets and tables)
        chunks = self._create_chunks(sheets_data, metadata)
        
        # Convert to markdown
        markdown_content = self._convert_to_markdown(sheets_data)
        
        return {
            'content': full_content,
            'markdown': markdown_content,
            'metadata': metadata,
            'chunks': chunks
        }

    def _process_sheet(self, sheet) -> Dict[str, Any]:
        """Process a single worksheet."""
        data = []
        max_col = 0
        used_rows = set()
        used_cols = set()
        
        # Find used cells
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    used_rows.add(cell.row)
                    used_cols.add(cell.column)
                    max_col = max(max_col, cell.column)
        
        if not used_rows or not used_cols:
            return {
                'title': sheet.title,
                'content': '',
                'rows': [],
                'metadata': {
                    'row_count': 0,
                    'column_count': 0,
                    'cell_count': 0,
                    'non_empty_cells': 0
                }
            }
        
        min_row, max_row = min(used_rows), max(used_rows)
        min_col, max_col = min(used_cols), max(used_cols)
        
        # Get headers (assuming first row contains headers)
        headers = []
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(min_row, col)
            headers.append(str(cell.value or f'Column {get_column_letter(col)}'))
        
        # Get data rows
        rows = []
        for row_idx in range(min_row + 1, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row_idx, col)
                row_data.append(self._format_cell_value(cell))
            rows.append(row_data)
        
        # Create markdown table content
        table_lines = []
        table_lines.append('| ' + ' | '.join(headers) + ' |')
        table_lines.append('| ' + ' | '.join(['---'] * len(headers)) + ' |')
        
        for row in rows:
            # Escape pipe characters and handle multiline content
            escaped_row = [
                str(cell).replace('|', '\\|').replace('\n', '<br>') 
                for cell in row
            ]
            table_lines.append('| ' + ' | '.join(escaped_row) + ' |')
        
        sheet_content = '\n'.join(table_lines)
        
        # Calculate sheet metadata
        non_empty_cells = sum(1 for row in rows for cell in row if cell is not None and str(cell).strip())
        
        return {
            'title': sheet.title,
            'content': sheet_content,
            'headers': headers,
            'rows': rows,
            'metadata': {
                'row_count': len(rows),
                'column_count': len(headers),
                'cell_count': len(rows) * len(headers),
                'non_empty_cells': non_empty_cells
            }
        }

    def _format_cell_value(self, cell) -> str:
        """Format cell value for display."""
        if cell.value is None:
            return ''
        
        # Handle different data types
        if cell.is_date:
            return cell.value.isoformat()[:10]
        elif isinstance(cell.value, (int, float)):
            if isinstance(cell.value, float) and cell.value.is_integer():
                return str(int(cell.value))
            return str(cell.value)
        
        return str(cell.value)

    def _extract_metadata(self, workbook, sheets_data: List[Dict]) -> Dict[str, Any]:
        """Extract metadata from the workbook."""
        total_rows = sum(sheet['metadata']['row_count'] for sheet in sheets_data)
        total_cells = sum(sheet['metadata']['cell_count'] for sheet in sheets_data)
        non_empty_cells = sum(sheet['metadata']['non_empty_cells'] for sheet in sheets_data)
        
        return {
            'sheet_count': len(workbook.sheetnames),
            'sheet_names': workbook.sheetnames,
            'total_rows': total_rows,
            'total_cells': total_cells,
            'non_empty_cells': non_empty_cells,
            'has_macros': workbook.vba_archive is not None,
            'properties': {
                'creator': workbook.properties.creator,
                'last_modified_by': workbook.properties.lastModifiedBy,
                'created': workbook.properties.created.isoformat() if workbook.properties.created else None,
                'modified': workbook.properties.modified.isoformat() if workbook.properties.modified else None,
            }
        }

    def _create_chunks(self, sheets_data: List[Dict], metadata: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Create chunks from sheets data."""
        chunks = []
        
        for sheet_data in sheets_data:
            if not sheet_data['rows']:
                continue
            
            # Create chunks of rows (50 rows per chunk)
            rows = sheet_data['rows']
            headers = sheet_data['headers']
            ROWS_PER_CHUNK = 50
            
            for i in range(0, len(rows), ROWS_PER_CHUNK):
                chunk_rows = rows[i:i + ROWS_PER_CHUNK]
                
                # Create markdown table for this chunk
                table_lines = []
                table_lines.append('| ' + ' | '.join(headers) + ' |')
                table_lines.append('| ' + ' | '.join(['---'] * len(headers)) + ' |')
                
                for row in chunk_rows:
                    escaped_row = [
                        str(cell).replace('|', '\\|').replace('\n', '<br>')
                        for cell in row
                    ]
                    table_lines.append('| ' + ' | '.join(escaped_row) + ' |')
                
                chunk_content = '\n'.join(table_lines)
                
                # Create chunk metadata
                chunk_metadata = {
                    'sheet_name': sheet_data['title'],
                    'row_range': {
                        'start': i + 1,
                        'end': min(i + ROWS_PER_CHUNK, len(rows))
                    },
                    'row_count': len(chunk_rows),
                    'column_count': len(headers),
                    'headers': headers
                }
                
                chunks.append({
                    'content': chunk_content,
                    'metadata': chunk_metadata
                })
        
        return chunks

    def _convert_to_markdown(self, sheets_data: List[Dict]) -> str:
        """Convert sheets data to markdown format."""
        markdown_parts = []
        
        for sheet_data in sheets_data:
            if sheet_data['content']:
                markdown_parts.append(f"## {sheet_data['title']}\n")
                markdown_parts.append(sheet_data['content'])
                markdown_parts.append('\n')  # Add space between sheets
        
        return '\n'.join(markdown_parts)
